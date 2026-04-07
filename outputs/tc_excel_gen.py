"""
Titan Carports — Construction Quote Excel Generator
Produces a professional .xlsx construction quote workbook.
"""

import io, datetime
import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# Brand colors (ARGB)
TC_DARK   = "FF1C1C2E"
TC_RED    = "FFC00000"
TC_BLUE   = "FF1F4E79"
TC_BLUE_M = "FF2E75B6"
TC_BLUE_L = "FFDEEAF1"
TC_GOLD   = "FFFFC000"
TC_WHITE  = "FFFFFFFF"
TC_LIGHT  = "FFF5F7FA"
TC_GREEN  = "FF375623"
TC_GOLD_L = "FFFFF8E1"

FMT_DOLLAR = '"$"#,##0.00'
FMT_INT    = '#,##0'


def fill(hex_argb):
    return PatternFill("solid", fgColor=hex_argb)

def font(bold=False, size=10, color="FF000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def border_thin():
    s = Side(style="thin", color="FFD0D7E2")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium_bottom():
    t = Side(style="thin", color="FFD0D7E2")
    m = Side(style="medium", color="FF2E75B6")
    return Border(left=t, right=t, top=t, bottom=m)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

def left_va():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_construction_quote_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Construction Quote"

    proj = data.get("project", {})
    sa   = data.get("sa", {})
    costs = data.get("costs", {})
    summary = data.get("summary", {})

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    row = 1

    # ── Header ────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, "TITAN CARPORTS — CONSTRUCTION QUOTE")
    c.font = font(bold=True, size=16, color=TC_WHITE)
    c.fill = fill(TC_DARK)
    c.alignment = center()
    row += 1

    ws.row_dimensions[row].height = 16
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, "Titan Carports · Conroe, TX")
    c.font = font(italic=True, size=9, color="FFA0A0A0")
    c.fill = fill(TC_DARK)
    c.alignment = center()
    row += 2

    # ── Project Info ──────────────────────────────────────────────────────
    info_pairs = [
        ("PROJECT", proj.get("name",""), "JOB CODE", proj.get("job_code","")),
        ("CUSTOMER", proj.get("customer_name",""), "QUOTE DATE", proj.get("quote_date","")),
        ("ADDRESS", f"{proj.get('address','')} {proj.get('city','')} {proj.get('state','')}",
         "MARKUP", f"{proj.get('markup_pct',35)}%"),
        ("SA QUOTE #", sa.get("quote_num",""), "", ""),
    ]
    for lbl1, val1, lbl2, val2 in info_pairs:
        ws.row_dimensions[row].height = 16
        c = ws.cell(row, 1, lbl1)
        c.font = font(bold=True, size=9, color=TC_BLUE)
        c.fill = fill(TC_BLUE_L)
        c.alignment = left_va()
        c.border = border_thin()
        c = ws.cell(row, 2, val1)
        c.font = font(size=9)
        c.alignment = left_va()
        c.border = border_thin()
        if lbl2:
            c = ws.cell(row, 3, lbl2)
            c.font = font(bold=True, size=9, color=TC_BLUE)
            c.fill = fill(TC_BLUE_L)
            c.alignment = left_va()
            c.border = border_thin()
            c = ws.cell(row, 4, val2)
            c.font = font(size=9)
            c.alignment = left_va()
            c.border = border_thin()
        row += 1
    row += 1

    # ── Cost Summary ──────────────────────────────────────────────────────
    # Section header
    ws.row_dimensions[row].height = 18
    c = ws.cell(row, 1, "COST CATEGORY")
    c.font = font(bold=True, size=10, color=TC_WHITE)
    c.fill = fill(TC_BLUE)
    c.alignment = center()
    c.border = border_thin()
    c = ws.cell(row, 2, "AMOUNT")
    c.font = font(bold=True, size=10, color=TC_WHITE)
    c.fill = fill(TC_BLUE)
    c.alignment = center()
    c.border = border_thin()
    row += 1

    materials_cost = float(sa.get("materials_cost", 0) or 0)
    sections_config = [
        ("Materials (SA Fabrication)", materials_cost),
        ("Concrete — Pier Footings",   float(costs.get("concrete",{}).get("total",0) or 0)),
        ("Labor — Installation",        float(costs.get("labor",{}).get("total",0) or 0)),
        ("Equipment Rental",            float(costs.get("equipment",{}).get("total",0) or 0)),
        ("Drilling",                    float(costs.get("drilling",{}).get("total",0) or 0)),
        ("Shipping & Freight",          float(costs.get("shipping",{}).get("total",0) or 0)),
        ("Fuel & Gas",                  float(costs.get("fuel",{}).get("total",0) or 0)),
        ("Hotels",                      float(costs.get("hotels",{}).get("total",0) or 0)),
        ("Per Diem",                    float(costs.get("per_diem",{}).get("total",0) or 0)),
        ("Transportation of Crew",      float(costs.get("transport",{}).get("total",0) or 0)),
        ("Miscellaneous",               float(costs.get("misc",{}).get("total",0) or 0)),
    ]

    subtotal = 0
    alt = True
    for label, amt in sections_config:
        if amt == 0:
            alt = not alt
            continue
        ws.row_dimensions[row].height = 16
        row_fill = fill("FFFFFFFF") if alt else fill(TC_LIGHT)
        c = ws.cell(row, 1, label)
        c.font = font(size=9)
        c.fill = row_fill
        c.alignment = left_va()
        c.border = border_thin()
        c2 = ws.cell(row, 2, amt)
        c2.font = font(size=9)
        c2.fill = row_fill
        c2.number_format = FMT_DOLLAR
        c2.alignment = right()
        c2.border = border_thin()
        subtotal += amt
        alt = not alt
        row += 1

    # Subtotal, Markup, Total
    markup_pct = float(proj.get("markup_pct", 35) or 35)
    markup_amt = subtotal * markup_pct / 100
    total = subtotal + markup_amt

    # Subtotal row
    ws.row_dimensions[row].height = 16
    c = ws.cell(row, 1, "SUBTOTAL")
    c.font = font(bold=True, size=9, color=TC_BLUE)
    c.fill = fill(TC_BLUE_L)
    c.alignment = left_va()
    c.border = border_medium_bottom()
    c = ws.cell(row, 2, subtotal)
    c.font = font(bold=True, size=9, color=TC_BLUE)
    c.fill = fill(TC_BLUE_L)
    c.number_format = FMT_DOLLAR
    c.alignment = right()
    c.border = border_medium_bottom()
    row += 1

    # Markup row
    ws.row_dimensions[row].height = 16
    c = ws.cell(row, 1, f"Markup ({markup_pct:.0f}%)")
    c.font = font(size=9)
    c.fill = fill(TC_GOLD_L)
    c.alignment = left_va()
    c.border = border_thin()
    c = ws.cell(row, 2, markup_amt)
    c.font = font(size=9)
    c.fill = fill(TC_GOLD_L)
    c.number_format = FMT_DOLLAR
    c.alignment = right()
    c.border = border_thin()
    row += 1

    # Total row
    ws.row_dimensions[row].height = 22
    c = ws.cell(row, 1, "TOTAL SELL PRICE")
    c.font = font(bold=True, size=13, color=TC_WHITE)
    c.fill = fill(TC_RED)
    c.alignment = left_va()
    c.border = border_thin()
    c = ws.cell(row, 2, total)
    c.font = font(bold=True, size=13, color=TC_WHITE)
    c.fill = fill(TC_RED)
    c.number_format = FMT_DOLLAR
    c.alignment = right()
    c.border = border_thin()
    row += 2

    # ── Equipment Detail Sheet ─────────────────────────────────────────────
    equip_items = costs.get("equipment", {}).get("items", [])
    if equip_items:
        ws2 = wb.create_sheet("Equipment Detail")
        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 8
        ws2.column_dimensions["C"].width = 10
        ws2.column_dimensions["D"].width = 14
        ws2.column_dimensions["E"].width = 14
        er = 1
        ws2.row_dimensions[er].height = 18
        hdr_cells = [("Description","A"), ("Qty","B"), ("Unit","C"), ("Rate","D"), ("Total","E")]
        for hdr, col_ltr in hdr_cells:
            c = ws2[f"{col_ltr}{er}"]
            c.value = hdr
            c.font = font(bold=True, size=9, color=TC_WHITE)
            c.fill = fill(TC_BLUE)
            c.alignment = center()
            c.border = border_thin()
        er += 1
        eq_total = 0
        alt2 = True
        for it in equip_items:
            qty = float(it.get("qty", 0) or 0)
            rate = float(it.get("rate", 0) or 0)
            line_total = qty * rate
            row_fill = fill("FFFFFFFF") if alt2 else fill(TC_LIGHT)
            row_vals = [it.get("desc",""), qty, it.get("unit","day"), rate, line_total]
            for ci, val in enumerate(row_vals, 1):
                c = ws2.cell(er, ci, val)
                c.font = font(size=9)
                c.fill = row_fill
                c.border = border_thin()
                if ci in (4, 5):
                    c.number_format = FMT_DOLLAR
                    c.alignment = right()
                elif ci == 2:
                    c.alignment = center()
            eq_total += line_total
            alt2 = not alt2
            er += 1
        # Total row
        c = ws2.cell(er, 4, "EQUIPMENT TOTAL")
        c.font = font(bold=True, size=9, color=TC_WHITE)
        c.fill = fill(TC_BLUE_M)
        c.alignment = right()
        c.border = border_thin()
        c = ws2.cell(er, 5, eq_total)
        c.font = font(bold=True, size=9, color=TC_WHITE)
        c.fill = fill(TC_BLUE_M)
        c.number_format = FMT_DOLLAR
        c.alignment = right()
        c.border = border_thin()

    # ── Footer on main sheet ───────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1,
        f"© {datetime.date.today().year} Titan Carports · Conroe, TX   "
        f"| Quote valid 30 days | Generated {datetime.date.today().strftime('%m/%d/%Y')}")
    c.font = font(italic=True, size=7, color="FF888888")
    c.alignment = center()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
