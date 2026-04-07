"""
Structures America — Excel BOM Generator
Produces a professional, branded Excel workbook with full material takeoff.
"""

import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from calc.defaults import COMPANY


# ─────────────────────────────────────────────
# COLORS (Structures America brand)
# ─────────────────────────────────────────────
SA_DARK_BG   = "1A1A2E"   # Dark navy background
SA_BLUE      = "1F4E79"   # Deep blue (headers)
SA_BLUE_MID  = "2E75B6"   # Mid blue (sub-headers)
SA_BLUE_LIGHT= "DEEAF1"   # Light blue (row alt)
SA_RED       = "C00000"   # Structures America red
SA_GOLD      = "FFC000"   # Accent gold
SA_GRAY_HDR  = "404040"   # Dark gray
SA_GRAY_ALT  = "F2F2F2"   # Alternating row
SA_WHITE     = "FFFFFF"
SA_GREEN     = "375623"   # Total rows
SA_GREEN_LIGHT = "E2EFDA"

def _border(style="thin"):
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    thick = Side(style="medium", color="404040")
    return Border(left=thick, right=thick, top=thick, bottom=thick)

def _cell(ws, row, col, value, bold=False, size=10, color=SA_GRAY_HDR,
          bg=None, align="left", wrap=False, fmt=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=size, color=color)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    ha = {"left": "left", "center": "center", "right": "right"}.get(align, "left")
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = border
    return c


# ─────────────────────────────────────────────
# MAIN GENERATOR
# ─────────────────────────────────────────────

def generate_bom_excel(bom_data: dict) -> bytes:
    """
    Generate Excel BOM workbook from bom_data dict.
    Returns bytes (xlsx file).
    """
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ── Summary Sheet ─────────────────────────
    _build_summary_sheet(wb, bom_data)

    # ── Per-Building BOM Sheets ───────────────
    for bldg in bom_data.get("buildings", []):
        _build_bom_sheet(wb, bldg, bom_data["project"])

    # ── Coil Summary Sheet ───────────────────
    if bom_data.get("summary_by_coil"):
        _build_coil_sheet(wb, bom_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_summary_sheet(wb: Workbook, bom_data: dict):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    col_widths = [3, 35, 15, 15, 15, 15, 3]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 50

    # ── Header ────────────────────────────────
    ws.merge_cells("B2:F2")
    c = ws.cell(row=2, column=2, value="STRUCTURES AMERICA")
    c.font = Font(name="Arial", bold=True, size=22, color=SA_WHITE)
    c.fill = PatternFill("solid", fgColor=SA_DARK_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Company info
    info_rows = [
        (3, COMPANY["address"]),
        (4, COMPANY["city_state_zip"]),
        (5, f"{COMPANY['contact_name']}  ·  {COMPANY['contact_phone']}"),
        (6, COMPANY["contact_email"]),
    ]
    for r, txt in info_rows:
        ws.row_dimensions[r].height = 15
        ws.merge_cells(f"B{r}:F{r}")
        c = ws.cell(row=r, column=2, value=txt)
        c.font = Font(name="Arial", size=9, color="666666")
        c.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[7].height = 6

    # ── Quote Info ────────────────────────────
    proj = bom_data["project"]
    ws.row_dimensions[8].height = 22
    ws.merge_cells("B8:F8")
    c = ws.cell(row=8, column=2, value="MATERIAL TAKEOFF & QUOTE")
    c.font = Font(name="Arial", bold=True, size=14, color=SA_WHITE)
    c.fill = PatternFill("solid", fgColor=SA_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")

    info_left = [
        ("Project Name:", proj.get("name", "")),
        ("Customer:", proj.get("customer_name", "")),
        ("Address:", f"{proj.get('address','')} {proj.get('city','')} {proj.get('state','')} {proj.get('zip_code','')}".strip()),
        ("Wind Speed:", f"{proj.get('wind_speed_mph', 115)} MPH"),
        ("Footing Depth:", str(proj.get('footing_depth_ft', 10)) + "' Deep"),
    ]
    info_right = [
        ("Quote #:", proj.get("job_code", "")),
        ("Quote Date:", proj.get("quote_date", datetime.date.today().strftime("%m/%d/%Y"))),
        ("Markup:", f"{proj.get('markup_pct', 35):.0f}%"),
        ("Prepared By:", COMPANY["contact_name"]),
        ("Contact:", COMPANY["contact_phone"]),
    ]

    for i, ((lbl, val), (lbl2, val2)) in enumerate(zip(info_left, info_right), 9):
        ws.row_dimensions[i].height = 16
        _cell(ws, i, 2, lbl, bold=True, size=9, color=SA_BLUE)
        _cell(ws, i, 3, val, size=9)
        _cell(ws, i, 4, "")
        _cell(ws, i, 5, lbl2, bold=True, size=9, color=SA_BLUE)
        _cell(ws, i, 6, val2, size=9, align="left")

    ws.row_dimensions[14].height = 8

    # ── Buildings Summary Table ───────────────
    headers = ["Building", "Type", "W×L (ft)", "# Bays", "Wt (LBS)", "Mat. Cost", "Fab Labor", "Sell Price"]
    hrow = 15
    ws.row_dimensions[hrow].height = 18
    for ci, h in enumerate(headers, 2):
        _cell(ws, hrow, ci, h, bold=True, size=9, color=SA_WHITE,
              bg=SA_BLUE, align="center", border=_border())

    row = hrow + 1
    total_wt = 0
    total_sell = 0
    total_labor = 0
    for bldg in bom_data.get("buildings", []):
        ws.row_dimensions[row].height = 15
        alt = SA_BLUE_LIGHT if (row % 2 == 0) else SA_WHITE
        labor_sell = bldg.get("labor_sell_price", 0.0)
        bldg_total = bldg.get("total_sell_price", 0) + labor_sell
        _cell(ws, row, 2, bldg.get("building_name", bldg.get("name", "Building")), size=9, bg=alt, border=_border())
        _cell(ws, row, 3, bldg["type"].upper(), size=9, bg=alt, align="center", border=_border())
        _cell(ws, row, 4, f"{bldg['width_ft']:.0f} × {bldg['length_ft']:.0f}", size=9, bg=alt, align="center", border=_border())
        _cell(ws, row, 5, bldg["geometry"].get("n_bays", ""), size=9, bg=alt, align="center", border=_border())
        _cell(ws, row, 6, bldg.get("total_weight_lbs", 0), size=9, bg=alt, align="right",
              fmt='#,##0', border=_border())
        _cell(ws, row, 7, bldg.get("total_material_cost", 0), size=9, bg=alt, align="right",
              fmt='"$"#,##0.00', border=_border())
        _cell(ws, row, 8, labor_sell if labor_sell > 0 else "—", size=9, bg=alt, align="right",
              fmt='"$"#,##0.00' if labor_sell > 0 else None, border=_border())
        _cell(ws, row, 9, bldg_total, size=9, bg=alt, align="right",
              fmt='"$"#,##0.00', border=_border())
        total_wt += bldg.get("total_weight_lbs", 0)
        total_sell += bldg_total
        total_labor += labor_sell
        row += 1

    # Total row
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"B{row}:E{row}")
    _cell(ws, row, 2, "TOTAL", bold=True, size=10, color=SA_WHITE, bg=SA_GREEN, align="right", border=_border())
    _cell(ws, row, 6, round(total_wt, 0), bold=True, size=10, color=SA_WHITE, bg=SA_GREEN, align="right", fmt='#,##0', border=_border())
    _cell(ws, row, 7, "", bold=True, bg=SA_GREEN, border=_border())
    _cell(ws, row, 8, round(total_labor, 2), bold=True, size=10, color=SA_WHITE, bg=SA_GREEN, align="right", fmt='"$"#,##0.00', border=_border())
    _cell(ws, row, 9, round(total_sell, 2), bold=True, size=10, color=SA_WHITE, bg=SA_GREEN, align="right", fmt='"$"#,##0.00', border=_border())

    row += 2

    # ── Signature Line ────────────────────────
    ws.row_dimensions[row].height = 40
    ws.merge_cells(f"B{row}:D{row}")
    c = ws.cell(row=row, column=2, value="Authorized Signature: ___________________________")
    c.font = Font(name="Arial", size=10, color="333333")
    c.alignment = Alignment(horizontal="left", vertical="bottom")

    ws.merge_cells(f"E{row}:F{row}")
    c = ws.cell(row=row, column=5, value="Date: ___________________________")
    c.font = Font(name="Arial", size=10, color="333333")
    c.alignment = Alignment(horizontal="left", vertical="bottom")

    row += 1
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"B{row}:F{row}")
    c = ws.cell(row=row, column=2,
                value=f"This quote is valid for 30 days. Prices subject to material market conditions. "
                      f"© {datetime.date.today().year} Structures America — Conroe, TX")
    c.font = Font(name="Arial", size=8, color="888888", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center")


def _build_bom_sheet(wb: Workbook, bldg: dict, proj: dict):
    bname = bldg.get("building_name", bldg.get("name", "Building"))
    name = bname[:28]
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    col_widths = [2, 28, 32, 10, 12, 10, 12, 12, 5, 36, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 26
    ws.merge_cells("B2:J2")
    c = ws.cell(row=2, column=2, value=f"BILL OF MATERIALS — {name.upper()}")
    c.font = Font(name="Arial", bold=True, size=14, color=SA_WHITE)
    c.fill = PatternFill("solid", fgColor=SA_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Geometry summary + fab time
    geo = bldg.get("geometry", {})
    labor = geo.get("labor", {})
    purlin_lbl = f"{geo.get('purlin_spacing_ft',5):.1f}' OC × {geo.get('n_purlin_lines','')} lines"
    purlin_lbl += " (auto)" if geo.get("purlin_auto") else " (override)"

    # Build fab time string for top of BOM
    fab_days = labor.get("total_fab_days", 0)
    fab_str = ""
    if fab_days > 0:
        fab_str = (
            f"Est. Fab: {fab_days} shop days  |  "
            f"Columns: {labor.get('days_columns',0)}d  "
            f"Rafters: {labor.get('days_rafters',0)}d  "
            f"Purlins: {labor.get('days_purlins',0)}d  "
            f"Panels: {labor.get('days_panels',0)}d  "
            f"Angle: {labor.get('days_angle',0)}d  "
            f"(4-man crew @ ${labor.get('daily_rate',960):.0f}/day + "
            f"{labor.get('overhead_pct',10):.0f}% overhead)"
        )

    geo_rows = [
        (f"Building: {bname}", f"Type: {bldg['type'].upper()}"),
        (f"Width: {bldg['width_ft']:.0f}' × Length: {bldg['length_ft']:.0f}'",
         f"Clear Height: {bldg['clear_height_ft']:.1f}'"),
        (f"Bays: {geo.get('n_bays','')} @ {geo.get('bay_size_ft',0):.2f}' ea  |  Overhang: {geo.get('overhang_ft',0):.2f}'",
         f"Purlin: {purlin_lbl}"),
        (f"Frames: {geo.get('n_frames','')} | Struct. Cols: {geo.get('n_struct_cols','')} | Rafters: {geo.get('n_rafters','')}",
         f"Slope: {geo.get('slope_deg',0):.2f}°  |  Col Ht: {geo.get('col_ht_ft',0):.2f}'"),
        (f"Rebar Col: {geo.get('rebar_col','—')} | Rebar Beam: {geo.get('rebar_beam','—')}",
         f"Quote #: {proj.get('job_code','')}"),
    ]
    for i, (l, r) in enumerate(geo_rows, 3):
        ws.row_dimensions[i+2].height = 14
        row = i + 2
        ws.merge_cells(f"B{row}:E{row}")
        _cell(ws, row, 2, l, size=9, color="444444")
        ws.merge_cells(f"F{row}:J{row}")
        _cell(ws, row, 6, r, size=9, color="444444")

    # Fabrication time banner row (row 8, between geo and headers)
    if fab_str:
        ws.row_dimensions[8].height = 16
        ws.merge_cells("B8:J8")
        c = ws.cell(row=8, column=2, value=fab_str)
        c.font = Font(name="Arial", bold=True, size=9, color=SA_WHITE)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Column headers (row 9 if no fab time, row 10 if fab time banner used)
    hrow = 10
    ws.row_dimensions[hrow].height = 20
    headers = [
        ("B", "Category", 28, "left"),
        ("C", "Description", 32, "left"),
        ("D", "Qty", 10, "right"),
        ("E", "Unit", 12, "center"),
        ("F", "Unit Wt\n(LBS)", 10, "right"),
        ("G", "Total Wt\n(LBS)", 12, "right"),
        ("H", "Unit Cost", 12, "right"),
        ("I", "Total Cost", 12, "right"),
        ("J", "Notes", 36, "left"),
    ]
    col_map = {"B":2,"C":3,"D":4,"E":5,"F":6,"G":7,"H":8,"I":9,"J":10}
    for col_letter, label, w, align in headers:
        ci = col_map[col_letter]
        _cell(ws, hrow, ci, label, bold=True, size=9, color=SA_WHITE,
              bg=SA_BLUE_MID, align=align, border=_border(), wrap=True)

    # Data rows
    row = hrow + 1
    current_cat = None
    for item in bldg.get("line_items", []):
        # Category header row
        if item["category"] != current_cat:
            current_cat = item["category"]
            ws.row_dimensions[row].height = 15
            ws.merge_cells(f"B{row}:J{row}")
            _cell(ws, row, 2, current_cat.upper(), bold=True, size=9,
                  color=SA_WHITE, bg=SA_BLUE, border=_border())
            row += 1

        ws.row_dimensions[row].height = 15
        alt = SA_GRAY_ALT if (row % 2 == 0) else SA_WHITE

        _cell(ws, row, 2, "", bg=alt, border=_border())
        _cell(ws, row, 3, item["description"], size=9, bg=alt, wrap=True, border=_border())
        _cell(ws, row, 4, item["qty"], size=9, bg=alt, align="right",
              fmt="#,##0.00", border=_border())
        _cell(ws, row, 5, item["unit"], size=9, bg=alt, align="center", border=_border())

        if item.get("unit_weight_lbs"):
            _cell(ws, row, 6, item["unit_weight_lbs"], size=9, bg=alt, align="right",
                  fmt="#,##0.000", border=_border())
        else:
            _cell(ws, row, 6, "", bg=alt, border=_border())

        if item.get("total_weight_lbs"):
            _cell(ws, row, 7, item["total_weight_lbs"], size=9, bg=alt, align="right",
                  fmt="#,##0.0", border=_border())
        else:
            _cell(ws, row, 7, "", bg=alt, border=_border())

        if item.get("unit_cost"):
            _cell(ws, row, 8, item["unit_cost"], size=9, bg=alt, align="right",
                  fmt='"$"#,##0.0000', border=_border())
        else:
            _cell(ws, row, 8, "", bg=alt, border=_border())

        if item.get("total_cost"):
            _cell(ws, row, 9, item["total_cost"], size=9, bg=alt, align="right",
                  fmt='"$"#,##0.00', border=_border())
        else:
            _cell(ws, row, 9, "", bg=alt, border=_border())

        _cell(ws, row, 10, item.get("notes", ""), size=8, bg=alt, wrap=True,
              color="555555", border=_border())
        row += 1

    # Totals
    row += 1
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"B{row}:F{row}")
    _cell(ws, row, 2, "TOTAL MATERIAL COST", bold=True, size=10, color=SA_WHITE,
          bg=SA_GREEN, align="right", border=_border())
    _cell(ws, row, 7, bldg.get("total_weight_lbs", 0), bold=True, size=10, color=SA_WHITE,
          bg=SA_GREEN, align="right", fmt="#,##0.0", border=_border())
    _cell(ws, row, 8, "", bg=SA_GREEN, border=_border())
    _cell(ws, row, 9, bldg.get("total_material_cost", 0), bold=True, size=10, color=SA_WHITE,
          bg=SA_GREEN, align="right", fmt='"$"#,##0.00', border=_border())
    _cell(ws, row, 10, "", bg=SA_GREEN, border=_border())

    # Labor line (if present)
    labor_sell = bldg.get("labor_sell_price", 0.0)
    labor_days = bldg.get("labor_total_days", 0)
    labor_raw  = bldg.get("labor_raw_cost", 0.0)
    if labor_sell > 0:
        row += 1
        ws.row_dimensions[row].height = 16
        ws.merge_cells(f"B{row}:F{row}")
        _cell(ws, row, 2,
              f"FABRICATION LABOR  ({labor_days} shop days | raw ${labor_raw:,.0f} + markup)",
              bold=True, size=9, color=SA_WHITE, bg="1F4E79", align="right", border=_border())
        _cell(ws, row, 7, "", bg="1F4E79", border=_border())
        _cell(ws, row, 8, "", bg="1F4E79", border=_border())
        _cell(ws, row, 9, labor_sell, bold=True, size=9, color=SA_WHITE,
              bg="1F4E79", align="right", fmt='"$"#,##0.00', border=_border())
        _cell(ws, row, 10, "", bg="1F4E79", border=_border())

    row += 1
    ws.row_dimensions[row].height = 20
    total_sell_all = bldg.get("total_sell_price", 0) + labor_sell
    ws.merge_cells(f"B{row}:H{row}")
    _cell(ws, row, 2, "SELL PRICE", bold=True, size=12,
          color=SA_WHITE, bg=SA_RED, align="right", border=_border())
    _cell(ws, row, 9, total_sell_all, bold=True, size=12, color=SA_WHITE,
          bg=SA_RED, align="right", fmt='"$"#,##0.00', border=_border())
    _cell(ws, row, 10, "", bg=SA_RED, border=_border())


def _build_coil_sheet(wb: Workbook, bom_data: dict):
    ws = wb.create_sheet("Coil Summary")
    ws.sheet_view.showGridLines = False

    col_widths = [2, 38, 14, 14, 14, 14, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 26
    ws.merge_cells("B2:F2")
    c = ws.cell(row=2, column=2, value="COIL MATERIAL SUMMARY")
    c.font = Font(name="Arial", bold=True, size=14, color=SA_WHITE)
    c.fill = PatternFill("solid", fgColor=SA_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")

    hrow = 4
    ws.row_dimensions[hrow].height = 18
    for ci, label in enumerate(["Material", "Total LFT", "Total LBS", "Unit Cost", "Total Cost"], 2):
        _cell(ws, hrow, ci, label, bold=True, size=9, color=SA_WHITE,
              bg=SA_BLUE_MID, align="center", border=_border())

    row = hrow + 1
    for cid, data in bom_data["summary_by_coil"].items():
        ws.row_dimensions[row].height = 15
        alt = SA_GRAY_ALT if (row % 2 == 0) else SA_WHITE
        _cell(ws, row, 2, data["description"], size=9, bg=alt, border=_border())
        _cell(ws, row, 3, data["total_lft"], size=9, bg=alt, align="right",
              fmt="#,##0.00", border=_border())
        _cell(ws, row, 4, round(data["total_lbs"], 1), size=9, bg=alt, align="right",
              fmt="#,##0.0", border=_border())
        _cell(ws, row, 6, round(data["total_cost"], 2), size=9, bg=alt, align="right",
              fmt='"$"#,##0.00', border=_border())
        row += 1

    row += 1
    ws.merge_cells(f"B{row}:E{row}")
    _cell(ws, row, 2, "PROJECT TOTALS", bold=True, size=10, color=SA_WHITE,
          bg=SA_GREEN, align="right", border=_border())
    _cell(ws, row, 6, bom_data.get("total_sell_price", 0), bold=True, size=10,
          color=SA_WHITE, bg=SA_GREEN, align="right", fmt='"$"#,##0.00', border=_border())
